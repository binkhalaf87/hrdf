from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from models import BankEmployee, BankReportRow, HadafEmployee, MatchResult, ProcessingSummary
from parser.bank_raw_extractor import BankRawRecord
from utils.logger import get_logger

logger = get_logger(__name__)

_GREEN  = "C6EFCE"
_YELLOW = "FFEB9C"
_RED    = "FFC7CE"
_BLUE   = "BDD7EE"
_ORANGE = "FCE4D6"
_HEADER_BG    = "2E75B6"
_HEADER_FONT  = "FFFFFF"


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = Font(bold=True, color=_HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _col_widths(ws, mn: int = 10, mx: int = 40) -> None:
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, mn), mx)


def _write_df(ws, df: pd.DataFrame, fill: str = "", rtl: bool = True) -> None:
    for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="right" if rtl else "left", vertical="center")
            if r > 1 and fill:
                cell.fill = PatternFill("solid", fgColor=fill)
    _style_header(ws)
    _col_widths(ws)
    if rtl:
        ws.sheet_view.rightToLeft = True


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ExcelWriter:

    # ------------------------------------------------------------------ #
    #  تقرير البنك المُحدَّث — الهدف الأساسي                             #
    #  الرقم التسلسلي هدف كأول عمود، ثم بيانات البنك كما هي             #
    # ------------------------------------------------------------------ #
    def build_bank_report_excel(self, rows: list[BankReportRow]) -> bytes:
        """
        ملف البنك مُحدَّث:
        - العمود الأول: رقم هدف التسلسلي (فارغ إذا لم يُطابَق)
        - باقي الأعمدة: بيانات البنك الأصلية
        - ألوان: أخضر=مطابق، أصفر=يحتاج مراجعة، أحمر=غير مطابق
        """
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("كشف الرواتب المُحدَّث")
        ws.sheet_view.rightToLeft = True

        # العمود الأول = رقم هدف، ثم بيانات البنك
        headers = [
            "رقم هدف",          # ← العمود المُضاف
            "اسم الموظف",
            "الآيبان",
            "المبلغ",
            "رقم المرجع",
        ]

        status_fill = {
            "matched":   _GREEN,
            "review":    _YELLOW,
            "bank_only": _RED,
        }

        # كتابة رأس الجدول
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.font = Font(bold=True, color=_HEADER_FONT)
            cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # كتابة البيانات — رقم هدف أولاً ثم باقي بيانات البنك
        for r_idx, brow in enumerate(rows, start=2):
            bg = PatternFill("solid", fgColor=status_fill.get(brow.status, "FFFFFF"))

            # رقم هدف التسلسلي (فارغ إذا لم يُطابَق، مع علامة ؟ للمراجعة)
            if brow.status == "matched":
                hadaf_serial_display = brow.hadaf_serial
            elif brow.status == "review":
                hadaf_serial_display = f"{brow.hadaf_serial}؟"   # يحتاج تأكيد
            else:
                hadaf_serial_display = ""   # غير مطابق — فارغ

            values = [
                hadaf_serial_display,
                brow.bank_name,
                brow.iban or "",
                brow.bank_amount,
                brow.reference or "",
            ]

            for c_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = bg
                cell.alignment = Alignment(horizontal="right", vertical="center")

        # تمييز عمود رقم هدف بلون مختلف للرأس
        ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1F4E79")
        ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF", size=12)

        _col_widths(ws, mn=8, mx=35)
        ws.column_dimensions["A"].width = 12  # عمود رقم هدف ثابت

        # إضافة ورقة تفصيلية للمطابقة
        ws2 = wb.create_sheet("تفاصيل المطابقة")
        ws2.sheet_view.rightToLeft = True
        detail_headers = [
            "رقم هدف",
            "اسم الموظف (هدف)",
            "اسم الموظف (البنك)",
            "الآيبان",
            "المبلغ (البنك)",
            "مبلغ هدف",
            "الفرق",
            "طريقة المطابقة",
            "نسبة الثقة %",
            "الحالة",
        ]
        for c_idx, h in enumerate(detail_headers, start=1):
            cell = ws2.cell(row=1, column=c_idx, value=h)
            cell.font = Font(bold=True, color=_HEADER_FONT)
            cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for r_idx, brow in enumerate(rows, start=2):
            bg = PatternFill("solid", fgColor=status_fill.get(brow.status, "FFFFFF"))
            detail_vals = [
                brow.hadaf_serial or "",
                brow.hadaf_name or "",
                brow.bank_name,
                brow.iban or "",
                brow.bank_amount,
                brow.hadaf_support_amount if brow.hadaf_support_amount is not None else "",
                brow.amount_diff if brow.amount_diff is not None else "",
                brow.match_method or "",
                f"{brow.confidence:.1f}%" if brow.confidence is not None else "",
                self._status_label(brow.status),
            ]
            for c_idx, val in enumerate(detail_vals, start=1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = bg
                cell.alignment = Alignment(horizontal="right", vertical="center")

        _col_widths(ws2, mn=10, mx=40)
        return _to_bytes(wb)

    # ------------------------------------------------------------------ #
    #  ملفات المطابقة التفصيلية                                           #
    # ------------------------------------------------------------------ #
    def build_matched_excel(self, results: list[MatchResult]) -> bytes:
        df = pd.DataFrame([{
            "الرقم التسلسلي (هدف)": r.hadaf_serial,
            "اسم هدف": r.hadaf_name,
            "اسم البنك": r.bank_name,
            "الآيبان": r.iban or "",
            "المبلغ (البنك)": r.bank_amount,
            "مبلغ هدف": r.hadaf_support_amount if r.hadaf_support_amount else "",
            "الفرق": r.amount_diff if r.amount_diff is not None else "",
            "طريقة المطابقة": r.match_method,
            "نسبة الثقة %": r.confidence,
        } for r in results])
        return self._single(df, "المطابقات", _GREEN)

    def build_review_excel(self, results: list[MatchResult]) -> bytes:
        df = pd.DataFrame([{
            "اسم البنك": r.bank_name,
            "الاسم المقترح (هدف)": r.hadaf_name,
            "الرقم التسلسلي المقترح": r.hadaf_serial,
            "الآيبان": r.iban or "",
            "المبلغ (البنك)": r.bank_amount,
            "مبلغ هدف": r.hadaf_support_amount if r.hadaf_support_amount else "",
            "الفرق": r.amount_diff if r.amount_diff is not None else "",
            "طريقة المطابقة": r.match_method,
            "نسبة الثقة %": r.confidence,
        } for r in results])
        return self._single(df, "للمراجعة", _YELLOW)

    def build_unmatched_excel(self, unmatched: list[BankEmployee]) -> bytes:
        df = pd.DataFrame([{
            "اسم البنك": e.name,
            "الآيبان": e.iban or "",
            "المبلغ": e.amount,
            "رقم المرجع": e.reference or "",
        } for e in unmatched])
        return self._single(df, "غير مطابق", _RED)

    def build_hadaf_not_in_bank_excel(self, employees: list[HadafEmployee]) -> bytes:
        """موظفو هدف الذين لم ينزل راتبهم."""
        df = pd.DataFrame([{
            "الرقم التسلسلي": e.serial,
            "اسم الموظف": e.name_arabic,
            "رقم الهوية": e.national_id or "",
            "الآيبان": e.iban or "",
            "مبلغ هدف": e.support_amount if e.support_amount else "",
        } for e in employees])
        return self._single(df, "هدف غير موجود بالبنك", _ORANGE)

    def build_hadaf_status_excel(
        self,
        employees: list[HadafEmployee],
        matched_serials: set[int],
        bank_serial_by_hadaf: dict[int, str] | None = None,
    ) -> bytes:
        """قائمة كاملة لموظفي هدف مع عمود حالة في البنك ورقم م من كشف البنك."""
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("موظفو هدف — حالة البنك")
        ws.sheet_view.rightToLeft = True

        has_bank_serial = bool(bank_serial_by_hadaf)
        headers = ["رقم هدف", "اسم الموظف", "رقم الهوية", "الآيبان", "مبلغ هدف", "حالة في البنك"]
        if has_bank_serial:
            headers.append("رقم م (البنك)")

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color=_HEADER_FONT)
            cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # تمييز عمود رقم م بلون مختلف
        if has_bank_serial:
            col_m = len(headers)
            ws.cell(row=1, column=col_m).fill = PatternFill("solid", fgColor="1F4E79")
        ws.row_dimensions[1].height = 28

        fill_green  = PatternFill("solid", fgColor=_GREEN)
        fill_orange = PatternFill("solid", fgColor=_ORANGE)

        for r, e in enumerate(sorted(employees, key=lambda x: x.serial), 2):
            in_bank = e.serial in matched_serials
            status  = "مضاف ✅" if in_bank else "غير مضاف ❌"
            fill    = fill_green if in_bank else fill_orange
            vals = [
                e.serial,
                e.name_arabic,
                e.national_id or "",
                e.iban or "",
                e.support_amount if e.support_amount is not None else "",
                status,
            ]
            if has_bank_serial:
                vals.append(bank_serial_by_hadaf.get(e.serial, ""))
            for c, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="right", vertical="center")

        _col_widths(ws, mn=10, mx=40)
        ws.column_dimensions["D"].width = 28  # IBAN
        ws.column_dimensions["F"].width = 16  # الحالة
        if has_bank_serial:
            ws.column_dimensions[chr(ord("F") + 1)].width = 14  # رقم م
        return _to_bytes(wb)

    def build_summary_excel(
        self,
        summary: ProcessingSummary,
        matched: list[MatchResult],
        review: list[MatchResult],
    ) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: ملخص أرقام
        ws = wb.create_sheet("ملخص")
        ws.sheet_view.rightToLeft = True
        rows = [
            ("البيان", "القيمة"),
            ("إجمالي موظفي هدف", summary.total_hadaf),
            ("إجمالي سجلات البنك", summary.total_bank),
            ("مطابق بنجاح ✅", summary.matched),
            ("تحتاج مراجعة ⚠️", summary.review_required),
            ("غير مطابق ❌", summary.unmatched),
            ("موظفو هدف لم ينزل راتبهم", summary.hadaf_not_in_bank),
            ("نسبة النجاح", f"{summary.success_rate:.1f}%"),
        ]
        fills_by_row = {
            1: _HEADER_BG,
            3: _GREEN, 4: _GREEN,
            5: _YELLOW,
            6: _RED,
            7: _ORANGE,
            8: _GREEN,
        }
        for r, (label, val) in enumerate(rows, start=1):
            ca = ws.cell(row=r, column=1, value=label)
            cb = ws.cell(row=r, column=2, value=val)
            bg = fills_by_row.get(r, "FFFFFF")
            for cell in (ca, cb):
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if r == 1:
                    cell.font = Font(bold=True, color=_HEADER_FONT)
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

        # Sheet 2: طرق المطابقة
        if matched or review:
            ws2 = wb.create_sheet("طرق المطابقة")
            ws2.sheet_view.rightToLeft = True
            all_r = matched + review
            df = (pd.DataFrame([{"method": r.match_method} for r in all_r])
                  .groupby("method").size().reset_index(name="count"))
            df.columns = ["طريقة المطابقة", "العدد"]
            _write_df(ws2, df, fill=_BLUE)

        return _to_bytes(wb)

    # ------------------------------------------------------------------ #
    #  Excel المطابقة بالآيبان — للنقل اليدوي                            #
    #  يحتوي فقط على السجلات المطابَقة (IBAN هدف = رقم الحساب البنكي)   #
    # ------------------------------------------------------------------ #
    def build_iban_matched_excel(
        self,
        bank_records: list[BankRawRecord],
        hadaf_by_iban: dict[str, int],          # IBAN.upper() → hadaf_serial
        hadaf_name_by_iban: dict[str, str],     # IBAN.upper() → Arabic name
        hadaf_amount_by_iban: dict[str, float], # IBAN.upper() → support amount
    ) -> bytes:
        """
        ينتج ورقتين:
          1. المطابَقون — السجلات التي تطابق فيها الآيبان فقط
          2. غير المطابَقين — موظفو هدف الذين لم يُعثر على آيبانهم في البنك
        """
        wb = Workbook()
        wb.remove(wb.active)

        # ── ورقة 1: المطابَقون ────────────────────────────────────────────
        ws = wb.create_sheet("مطابَق بالآيبان")
        ws.sheet_view.rightToLeft = True

        headers = [
            "رقم هدف",          # A — العمود الأساسي المُضاف
            "م (بنك)",          # B — رقم تسلسلي البنك
            "اسم الموظف (هدف)", # C — الاسم العربي من ملف هدف
            "اسم الموظف (بنك)", # D — الاسم الإنجليزي من البنك
            "الآيبان",          # E
            "رقم المرجع",       # F
            "رمز البنك",        # G
            "المبلغ (بنك)",     # H
            "مبلغ هدف",         # I
        ]

        # Header row style
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color=_HEADER_FONT)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # رأس عمود رقم هدف بلون مميز
        ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1F4E79")
        # باقي الرأس
        for c in range(2, len(headers) + 1):
            ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=_HEADER_BG)

        matched_rows = []
        for rec in bank_records:
            iban_up = rec.iban.upper() if rec.iban else ""
            if iban_up not in hadaf_by_iban:
                continue
            matched_rows.append({
                "hadaf_serial":  hadaf_by_iban[iban_up],
                "bank_serial":   rec.bank_serial,
                "hadaf_name":    hadaf_name_by_iban.get(iban_up, ""),
                "bank_name":     rec.name,
                "iban":          rec.iban,
                "reference":     rec.reference,
                "bank_code":     rec.bank_code,
                "bank_amount":   rec.amount_str,
                "hadaf_amount":  hadaf_amount_by_iban.get(iban_up, ""),
            })

        # Sort by Hadaf serial
        matched_rows.sort(key=lambda x: x["hadaf_serial"])

        fill_green = PatternFill("solid", fgColor=_GREEN)
        for r, row in enumerate(matched_rows, 2):
            vals = [
                row["hadaf_serial"],
                row["bank_serial"],
                row["hadaf_name"],
                row["bank_name"],
                row["iban"],
                row["reference"],
                row["bank_code"],
                row["bank_amount"],
                row["hadaf_amount"] if row["hadaf_amount"] else "",
            ]
            for c, val in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.fill = fill_green
                cell.alignment = Alignment(horizontal="right", vertical="center")

        _col_widths(ws, mn=10, mx=40)
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["E"].width = 28   # IBAN
        ws.column_dimensions["F"].width = 20   # reference
        ws.row_dimensions[1].height = 28

        # ── ورقة 2: غير المطابَقين من هدف ───────────────────────────────
        ws2 = wb.create_sheet("هدف بدون تطابق بنكي")
        ws2.sheet_view.rightToLeft = True

        bank_ibans = {rec.iban.upper() for rec in bank_records if rec.iban}
        unmatched_hadaf = [
            iban_up for iban_up in hadaf_by_iban
            if iban_up not in bank_ibans
        ]

        h2 = ["رقم هدف", "اسم الموظف", "الآيبان"]
        for c, h in enumerate(h2, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color=_HEADER_FONT)
            cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        fill_red = PatternFill("solid", fgColor=_RED)
        for r, iban_up in enumerate(sorted(unmatched_hadaf,
                                           key=lambda x: hadaf_by_iban[x]), 2):
            vals = [hadaf_by_iban[iban_up], hadaf_name_by_iban.get(iban_up, ""), iban_up]
            for c, val in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=c, value=val)
                cell.fill = fill_red
                cell.alignment = Alignment(horizontal="right", vertical="center")

        _col_widths(ws2, mn=10, mx=35)

        return _to_bytes(wb)

    @staticmethod
    def _status_label(status: str) -> str:
        return {"matched": "مطابق ✅", "review": "يحتاج مراجعة ⚠️",
                "bank_only": "غير مطابق ❌"}.get(status, status)

    @staticmethod
    def _single(df: pd.DataFrame, sheet: str, fill: str) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(sheet)
        _write_df(ws, df, fill=fill)
        return _to_bytes(wb)
